#!/usr/bin/env python
# coding: utf-8

# In[21]:


import pandas as pd
import random
from openpyxl import Workbook
from openpyxl import load_workbook
import math
from scipy import stats


# In[35]:


# leia o arquivo do excel em um dataframe do pandas
df = pd.read_excel('trfrj1Amostra.xlsx')


# In[36]:


# defina o nível de confiança e a margem de erro desejados
confianca = 0.95
margem_erro = 0.05

# Determine o valor crítico da distribuição normal para o nível de confiança desejado
z = abs(stats.norm.ppf((1 - confianca) / 2))

# Estime a proporção da característica de interesse na população (pode ser 0,5 para maximizar o tamanho da amostra)
p = 0.5

# obtenha o tamanho da população
tamanho_populacao = len(df) - 1

# calcule o tamanho da amostra necessário com base na fórmula do tamanho da amostra
tamanho_amostra = math.ceil((tamanho_populacao * z**2 * p * (1-p)) / ((tamanho_populacao - 1) * margem_erro**2 + z**2 * p * (1-p)))


# In[37]:


# selecione uma amostra aleatória de linhas
amostra = df.sample(n=tamanho_amostra, random_state=1)

# salve a amostra como um arquivo do Excel
amostra.to_excel('trfrj1Amostra-Selecionados.xlsx', sheet_name='Amostra', index=False)


# In[38]:


# atualize o arquivo original para indicar quais linhas foram selecionadas
wb_orig = load_workbook(filename="trfrj1Amostra.xlsx")
ws_orig = wb_orig.active

# crie uma nova coluna para indicar se a linha foi selecionada ou não
ws_orig.insert_cols(len(df.columns) + 1)
ws_orig.cell(row=1, column=len(df.columns) + 1, value='Selecionado')

# preencha a nova coluna com os valores 'Sim' ou 'Não' para cada linha
for row_num, row_data in df.iterrows():
    if row_num + 2 in amostra.index:
        ws_orig.cell(row=row_num + 2, column=len(df.columns) + 1, value='Sim')
    else:
        ws_orig.cell(row=row_num + 2, column=len(df.columns) + 1, value='Não')

# salve o arquivo original atualizado
wb_orig.save('trfrj1Amostra_atualizado.xlsx')


# In[ ]:




